from __future__ import annotations

import hashlib
import io
import re
import zipfile
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

from .models import WBWeeklyReference


_DATE_RANGE_RE = re.compile(r"(\d{2}\.\d{2}\.\d{4})\s*[-—–]\s*(\d{2}\.\d{2}\.\d{4})")
_REPORT_NUMBER_RE = re.compile(r"(?:№|N)\s*(\d{4,})", re.IGNORECASE)

_FIELD_ALIASES = {
    "revenue": ("выручка", "продажа", "реализация"),
    "payout": ("к перечислению продавцу", "к перечислению за товар", "итого к оплате", "к оплате"),
    "logistics": ("логистика",),
    "storage": ("хранение",),
    "acquiring": ("эквайринг",),
    "wb_deductions": ("удержания",),
    "other_expenses": ("прочие удержания", "прочие расходы"),
    "penalties": ("штрафы",),
    "advertising": ("реклама", "реклам", "продвижение"),
    "orders_count": ("количество продаж", "количество заказов", "заказы"),
    "buyouts_count": ("количество выкупов", "выкупы", "продажи шт"),
    "returns_count": ("количество возвратов", "возвраты", "возвратов"),
}


def _normalize_text(value: Any) -> str:
    return str(value or "").strip().lower().replace("\n", " ")


def _number(value: Any) -> float | None:
    if value in (None, ""):
        return None
    if isinstance(value, (int, float)):
        return round(float(value), 2)
    text = str(value).strip().replace("\xa0", " ").replace("₽", "").replace("р.", "")
    text = text.replace(" ", "").replace(",", ".")
    match = re.search(r"-?\d+(?:\.\d+)?", text)
    if not match:
        return None
    try:
        return round(float(match.group(0)), 2)
    except Exception:
        return None


def _count_number(value: Any) -> int | None:
    amount = _number(value)
    if amount is None:
        return None
    return int(round(amount))


def _parse_date(text: str) -> date:
    return datetime.strptime(text, "%d.%m.%Y").date()


def _find_period(texts: list[str], file_name: str) -> tuple[date, date]:
    for text in [file_name, *texts]:
        match = _DATE_RANGE_RE.search(str(text))
        if match:
            return _parse_date(match.group(1)), _parse_date(match.group(2))
    raise ValueError("WB weekly report period not found")


def _find_report_number(texts: list[str], file_name: str) -> str | None:
    for text in [file_name, *texts]:
        match = _REPORT_NUMBER_RE.search(str(text))
        if match:
            return match.group(1)
    return None


def _sheet_rows_from_bytes(raw_bytes: bytes) -> tuple[list[tuple[str, list[list[Any]]]], list[str]]:
    workbook = load_workbook(io.BytesIO(raw_bytes), read_only=True, data_only=True)
    try:
        sheets: list[tuple[str, list[list[Any]]]] = []
        texts: list[str] = []
        for sheet in workbook.worksheets:
            rows = [list(row) for row in sheet.iter_rows(values_only=True)]
            sheets.append((sheet.title, rows))
            for row in rows:
                texts.extend([str(cell) for cell in row if cell not in (None, "")])
        return sheets, texts
    finally:
        workbook.close()


def _extract_totals(sheets: list[tuple[str, list[list[Any]]]]) -> tuple[dict[str, Any], dict[str, Any]]:
    found: dict[str, Any] = {}
    raw_totals: dict[str, Any] = {}
    for sheet_name, rows in sheets:
        for row in rows:
            normalized_cells = [_normalize_text(cell) for cell in row]
            for field_name, aliases in _FIELD_ALIASES.items():
                if field_name in found:
                    continue
                hit_index = None
                for index, cell in enumerate(normalized_cells):
                    if any(alias in cell for alias in aliases):
                        hit_index = index
                        break
                if hit_index is None:
                    continue
                numeric_value = None
                for cell in row[hit_index + 1 :]:
                    numeric_value = _number(cell)
                    if numeric_value is not None:
                        break
                if numeric_value is None:
                    for cell in reversed(row[:hit_index]):
                        numeric_value = _number(cell)
                        if numeric_value is not None:
                            break
                if numeric_value is None:
                    continue
                parsed_value: Any = _count_number(numeric_value) if field_name.endswith("_count") else float(numeric_value)
                found[field_name] = parsed_value
                raw_totals[field_name] = {
                    "value": parsed_value,
                    "sheet": sheet_name,
                    "row": [cell for cell in row],
                }
    return found, raw_totals


def _read_reference_file(file_path: str) -> tuple[str, bytes, bytes]:
    path = Path(file_path)
    outer_bytes = path.read_bytes()
    if path.suffix.lower() != ".zip":
        return path.name, outer_bytes, outer_bytes
    with zipfile.ZipFile(io.BytesIO(outer_bytes), "r") as archive:
        workbook_members = [name for name in archive.namelist() if Path(name).suffix.lower() in (".xlsx", ".xlsm")]
        if not workbook_members:
            raise ValueError("ZIP does not contain .xlsx workbook")
        member_name = workbook_members[0]
        return member_name, archive.read(member_name), outer_bytes


def load_wb_weekly_reference(file_path: str) -> WBWeeklyReference:
    workbook_name, workbook_bytes, source_bytes = _read_reference_file(file_path)
    sheets, texts = _sheet_rows_from_bytes(workbook_bytes)
    period_from, period_to = _find_period(texts, Path(file_path).name)
    report_number = _find_report_number(texts, Path(file_path).name)
    totals, raw_totals = _extract_totals(sheets)
    return WBWeeklyReference(
        source_file=str(file_path),
        source_hash=hashlib.sha256(source_bytes).hexdigest(),
        period_from=period_from,
        period_to=period_to,
        report_number=report_number,
        revenue=totals.get("revenue"),
        payout=totals.get("payout"),
        logistics=totals.get("logistics"),
        storage=totals.get("storage"),
        acquiring=totals.get("acquiring"),
        wb_deductions=totals.get("wb_deductions"),
        other_expenses=totals.get("other_expenses"),
        penalties=totals.get("penalties"),
        advertising=totals.get("advertising"),
        orders_count=totals.get("orders_count"),
        buyouts_count=totals.get("buyouts_count"),
        returns_count=totals.get("returns_count"),
        raw_totals=raw_totals,
        metadata={
            "workbook_name": workbook_name,
            "sheet_names": [name for name, _rows in sheets],
            "source_kind": "zip" if str(file_path).lower().endswith(".zip") else "xlsx",
        },
    )
