import os
from datetime import datetime

from openpyxl import Workbook

from report import (
    get_advertising_stats,
    get_orders_stats,
    get_profit_stats,
    get_replenishment_plan,
    get_top_profit_products,
    normalize_period,
)

EXPORT_DIR = 'exports'


def _ensure():
    os.makedirs(EXPORT_DIR, exist_ok=True)


def export_pnl(telegram_id, period='month'):
    _ensure()
    pname, days = normalize_period(period)
    wb = Workbook()
    ws = wb.active
    ws.title = 'P&L'
    st = get_profit_stats(days, telegram_id)
    od = get_orders_stats(days, telegram_id)
    ad = get_advertising_stats(days, telegram_id)
    rows = [
        ('Период', pname),
        ('Заказы', od[0]),
        ('Сумма заказов', od[1]),
        ('Отмены', od[2]),
        ('Выручка продаж', st[0]),
        ('Комиссия WB', st[1]),
        ('К выплате', st[2]),
        ('Себестоимость', st[3]),
        ('Логистика', st[4]),
        ('Реклама', st[5]),
        ('Хранение', st[6]),
        ('Прочее', st[7]),
        ('Расходы всего', st[8]),
        ('Прибыль до расходов', st[9]),
        ('Чистая прибыль', st[10]),
        ('Маржинальность %', st[11]),
        ('Возвраты', st[12]),
        ('Показы рекламы', ad[0]),
        ('Клики', ad[1]),
        ('ROAS', ad[7]),
        ('ДРР %', ad[8]),
    ]
    for row in rows:
        ws.append(row)
    path = os.path.join(EXPORT_DIR, f"pnl_{telegram_id}_{pname}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb.save(path)
    return path


def export_products(telegram_id, period='month'):
    _ensure()
    pname, days = normalize_period(period)
    wb = Workbook()
    ws = wb.active
    ws.title = 'Products'
    ws.append([
        'Артикул', 'Выплата', 'Себестоимость', 'Прибыль до расходов', 'Логистика',
        'Реклама', 'Хранение', 'Прочее', 'Расходы', 'Чистая прибыль', 'Маржа %', 'Возвраты'
    ])
    for row in get_top_profit_products(days=days, telegram_id=telegram_id, limit=1000):
        ws.append(list(row))
    path = os.path.join(EXPORT_DIR, f"products_{telegram_id}_{pname}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb.save(path)
    return path


def export_replenishment(telegram_id):
    _ensure()
    plan = get_replenishment_plan(telegram_id)
    settings = plan['settings']
    wb = Workbook()
    ws = wb.active
    ws.title = 'Replenishment'
    ws.append(['Параметр', 'Значение'])
    ws.append(['Окно продаж, дней', settings['sales_window_days']])
    ws.append(['Целевой запас, дней', settings['target_stock_days']])
    ws.append(['Срок поставки, дней', settings['lead_time_days']])
    ws.append(['Страховой запас, дней', settings['safety_stock_days']])
    ws.append(['Минимальная партия', settings['min_order_qty']])
    ws.append(['Суммарно к заказу, шт', plan['total_units']])
    ws.append(['Бюджет закупки', plan['total_budget']])
    ws.append([])
    ws.append([
        'Артикул', 'Остаток', 'Всего на складах', 'В пути к клиенту', 'В пути от клиента',
        'Средние продажи в день', 'Дней до OOS', 'Дата OOS', 'Риск', 'Рекомендуемый заказ', 'Бюджет', 'cost_missing'
    ])
    for row in plan['items']:
        ws.append([
            row['article'],
            row['current_stock'],
            row['quantity_full'],
            row['in_way_to_client'],
            row['in_way_from_client'],
            row['avg_sales_per_day'],
            row['days_left'],
            row['oos_date'],
            row['risk_level'],
            row['need'],
            row['budget'],
            'Да' if row.get('cost_missing') else 'Нет',
        ])
    path = os.path.join(EXPORT_DIR, f"replenishment_{telegram_id}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx")
    wb.save(path)
    return path
